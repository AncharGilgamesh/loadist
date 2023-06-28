"""Class for work with Excel file."""


from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Alignment
import Models.DataHandler as DH
import Models.SqlHandler as SQH
import os
import xlrd


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SHABLON_PATH = os.path.join(
    BASE_DIR, '..', 'Resources', 'Shablons', 'shablon.xlsx')
SHABLON_NAGRUZKA_PATH = os.path.join(
    BASE_DIR, '..', 'Resources', 'Shablons', 'nagruzka_shablon.xlsm'
)


class ExcelHandler:
    """Class."""

    def __init__(self, filename=None):
        """func."""
        self._filename = filename
        self.workbook = None
        self.sheet = None

    def open_workbook(self):
        """func."""
        self.workbook = load_workbook(filename=self._filename, data_only=True)

    def create_workbook(self):
        """func."""

    def get_sheet_by_name(self, name):
        """func."""
        self.sheet = self.workbook[name]

    def get_sheet_names(self):
        """func."""
        return self.workbook.sheetnames

    def create_sheet(self, name):
        """func."""
        self.sheet = self.workbook.create_sheet(name)

    def get_cell_value(self, row, column):
        """func."""
        return self.sheet.cell(row=row, column=column).value

    def set_cell_value(self, row, column, value):
        """func."""
        self.sheet.cell(row=row, column=column, value=value)

    def save_workbook(self):
        """func."""
        self.workbook.save(filename=self.filename)

    def read_rows_to_list(self):
        """
        func.

        Функция для записи строк таблицы Excel в переменную
        """
        headers = []
        avoid_indexes = []
        data = []
        sheet_borders = SQH.select_settings()[0][-4:]
        min_row = 3
        max_row = self.sheet.max_row
        min_col = 1
        max_col = 27
        if sheet_borders[0]:
            min_row = sheet_borders[0]
        if sheet_borders[1]:
            max_row = sheet_borders[1]
        if sheet_borders[2]:
            min_col = sheet_borders[2]
        if sheet_borders[3]:
            max_col = sheet_borders[3]
        for cell in self.sheet[2]:
            """
            Тут начинается с единицы счет, поэтому индексы на + 1 в enumerate
            """
            if 'всего' in str(cell.value):
                break
            if str(cell.value) == '1':
                avoid_indexes.append(cell.column)
                continue
            if str(cell.value) is None:
                avoid_indexes.append(cell.column)
                continue
            if cell.data_type == 'n':
                avoid_indexes.append(cell.column)
                continue
            headers.append(cell.value)
        for row in self.sheet.iter_rows(
         min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col,
         values_only=True):
            if 'Итого' in row:
                break
            row_values = []
            for i, cell in enumerate(row):
                if i + 1 in avoid_indexes:
                    continue
                row_values.append(cell)
            data.append(row_values)
        return headers, data

    def write_nagruzka(self, data, savePath):
        """func."""
        start_row = 3
        source_wb = load_workbook(SHABLON_NAGRUZKA_PATH)
        source_sheet = source_wb['нагрузка кафедры']
        self.fill_nagruzka_poles(source_sheet, data)
        self.fillSemestr(source_sheet, data, start_row, 1)
        self.addBorders(source_sheet, start_row, len(data) + 2)
        self.sumVsego(source_sheet, start_row, len(data) + 3)
        sumlist = self.summasCol(3, len(data) + 2)
        self.addItogoSumRow(sumlist, source_sheet, len(data) + 3)
        self.sum_all_nagruzka(source_sheet, 3, len(data) + 2)
        source_wb.save(savePath)

    def write_statistics(self, data, save_path):
        """func."""
        wb = Workbook()
        sheet = wb.create_sheet()
        self.fillSemestr(sheet, data, 5, 3)
        self.addBorders(sheet, 5, len(data) + 4, len(data[0]) + 2, 3)
        wb.save(save_path)

    def writeRows(self, osen, vesna, savePath, tchr_id):
        """func."""
        rowstoInsertOsen = len(osen)
        rowstoInsertVesna = len(vesna)
        source_wb = load_workbook(SHABLON_PATH)
        source_sheet = source_wb['ШАА']
        osenRow, vesnaRow, itogoRow = self.findRows(
            source_sheet, rowstoInsertOsen, rowstoInsertVesna)
        start_rowOsen = osenRow + 1
        start_rowVesna = vesnaRow + 1

        self.fillSemestr(source_sheet, osen, start_rowOsen, 1)
        self.fillSemestr(source_sheet, vesna, start_rowVesna, 1)
        self.addBorders(source_sheet, osenRow, vesnaRow)
        self.addBorders(source_sheet, vesnaRow, itogoRow)

        self.sumVsego(source_sheet, osenRow + 1, vesnaRow - 1)
        self.sumVsego(source_sheet, vesnaRow + 1, itogoRow - 1)

        self.addDopColSum(source_sheet, osenRow + 1, vesnaRow - 1, osenRow)
        self.addDopColSum(source_sheet, vesnaRow + 1, itogoRow - 1, vesnaRow)

        self.addItogSemSum(source_sheet, osenRow + 1, vesnaRow - 1, osenRow)
        self.addItogSemSum(source_sheet, vesnaRow + 1, itogoRow - 1, vesnaRow)

        self.addAllSemSum(source_sheet, osenRow, vesnaRow, itogoRow)
        sumlist = self.summasCol(8, itogoRow - 1)
        self.addItogoSumRow(sumlist, source_sheet, itogoRow)
        self.addDopColAll(source_sheet, osenRow, vesnaRow, itogoRow)
        self.fill_teacher_info(source_sheet, tchr_id)
        source_wb.save(savePath)

    def addDopColAll(self, sheet, osen, vesna, rowNum):
        """func."""
        row = sheet[rowNum]
        val = f'=SUM(AA{osen} + AA{vesna})'
        row[26].value = val

    def sum_all_nagruzka(self, sheet, start, end):
        """func."""
        row = sheet[end + 1]
        val = f'=SUM(AB{start}:AB{end})'
        row[27].value = val

    def summasCol(self, start, end):
        """func."""
        L = f'=SUM(L{start}:L{end})'
        M = f'=SUM(M{start}:M{end})'
        N = f'=SUM(N{start}:N{end})'
        Oo = f'=SUM(O{start}:O{end})'
        P = f'=SUM(P{start}:P{end})'
        Q = f'=SUM(Q{start}:Q{end})'
        R = f'=SUM(R{start}:R{end})'
        S = f'=SUM(S{start}:S{end})'
        T = f'=SUM(T{start}:T{end})'
        U = f'=SUM(U{start}:U{end})'
        V = f'=SUM(V{start}:V{end})'
        W = f'=SUM(W{start}:W{end})'
        X = f'=SUM(X{start}:X{end})'
        Y = f'=SUM(Y{start}:Y{end})'
        Z = f'=SUM(Z{start}:Z{end})'
        AA = f'=SUM(AA{start}:AA{end})'
        return [L, M, N, Oo, P, Q, R, S, T, U, V, W, X,
                Y, Z, AA]

    def sumRow(self, start):
        """func."""
        return f'=SUM(L{start}:AA{start})'

    def sumVsego(self, sheet, start, end):
        """func."""
        for i in range(start, end):
            val = self.sumRow(i)
            sheet.cell(row=i, column=28).value = val

    def fill_nagruzka_poles(self, sheet, data):
        """func."""
        sheet.insert_rows(3, len(data) - 1)

    def findRows(self, sheet, insertOsen, insertVesna):
        """func."""
        i = 1
        osenRow = 0
        vesnaRow = 0
        itogoRow = 0
        while True:
            if sheet.cell(row=i, column=3).value == 'Осенний семестр':
                osenRow = i
                if insertOsen:
                    # Добавляем строки только в том случае, если есть что
                    #  добавить
                    sheet.insert_rows(8, insertOsen)
            if sheet.cell(row=i, column=3).value == 'Весенний семестр':
                vesnaRow = i
                if insertVesna:
                    sheet.insert_rows(vesnaRow+1, insertVesna)
            if sheet.cell(row=i, column=3).value == 'Итого за год':
                itogoRow = i
                break
            i += 1
        return osenRow, vesnaRow, itogoRow

    def fillSemestr(self, sheet, data, start_row, start_col):
        """Func."""
        for row_index, row in enumerate(data):
            for col_index, value in enumerate(row):
                # Получаем ячейку по индексам строки и столбца
                cell = sheet.cell(
                    row=start_row+row_index, column=start_col+col_index)
                # Записываем значение в ячейку
                cell.value = value

    def addBorders(self, sheet, start_row, end_row, max_col=28, min_col=1):
        """func."""
        thin_border = Border(left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))
        for row in sheet.iter_rows(
             min_row=start_row,
             max_col=max_col, max_row=end_row, min_col=min_col):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

    def addItogoSumRow(self, sumlist, sheet, rowNum, start_col=12):
        """func."""
        row = sheet[rowNum]
        for i, val in enumerate(sumlist):
            row[start_col+i-1].value = val

    def addDopColSum(self, sheet, start, end, rowNum):
        """func."""
        """
        row - куда вставляем
        """
        row = sheet[rowNum]
        val = f'=SUM(AA{start}:AA{end})'
        row[26].value = val

    def addItogSemSum(self, sheet, start, end, rowNum):
        """func."""
        row = sheet[rowNum]
        val = f'=SUM(AB{start}:AB{end})'
        row[27].value = val

    def addAllSemSum(self, sheet, osen, vesna, rowNum):
        """func."""
        row = sheet[rowNum]
        val = f'=SUM(AB{osen} + AB{vesna})'
        row[27].value = val
        sheet['E4'] = val

    def fill_teacher_info(self, sheet, tbl_id):
        """func."""
        prepod_info = DH.get_prepod_info(tbl_id)
        setting_info = SQH.select_settings()
        setting_info = tuple(map(str, setting_info[0]))
        study_year = 'Индивидуальный план на ' + setting_info[0] + '/'
        study_year += setting_info[1] + ' уч.год'
        sheet['E1'] = study_year
        sheet['E2'] = setting_info[2]
        sheet['K2'] = setting_info[3]
        # sheet['E2'] = 'информационных технологий и энергетических систем'
        # sheet['K2'] = 'системного  анализа и информатики'
        # name
        sheet['E3'] = prepod_info[0]
        # zvanie
        if prepod_info[1]:
            sheet['R3'] = 'Ученое звание: ' + prepod_info[1]
        # stepen
        if prepod_info[2]:
            sheet['M3'] = 'Ученая степень: ' + prepod_info[2]
        # doljnost
        sheet['I3'] = 'Должность: ' + prepod_info[3]
        # stavka
        if not prepod_info[5]:
            sheet['L4'] = prepod_info[4]
        # pochasovka
        if not prepod_info[6]:
            sheet['AA4'] = prepod_info[5]
        # pochasovka v osn(dogovora)
        if not prepod_info[5]:
            sheet['AA4'] = prepod_info[6]
        # sovmestitelstvo
        sheet['R4'] = prepod_info[7]
        # stavka po doljnosti
        if not prepod_info[5]:
            sheet['AA3'] = prepod_info[8]


class XlsHandler:
    """class."""

    def __init__(self, filename):
        """func."""
        self._filename = filename
        self._workbook = None
        self._sheet = None
        self._nagruzka_names = [
            'нагрузка кафедры',
            'нагрузка',
            'общая'
        ]

    def open_workbook(self):
        """func."""
        self._workbook = xlrd.open_workbook(self._filename)

    def get_nagruzka_sheet(self):
        """func."""
        sheet_names = self._workbook.sheet_names()
        for sheet_name in sheet_names:
            if sheet_name in self._nagruzka_names:
                self._sheet = self._workbook.sheet_by_name(sheet_name)
                return
        return None

    def read_rows_to_list(self):
        """func."""
        # workbook = xlrd.open_workbook(file_path)
        sheet = self._sheet
        headers = []
        avoid_indexes = []
        data = []
        sheet_borders = SQH.select_settings()[0][-4:]
        min_row = 2
        max_row = sheet.nrows
        min_col = 0
        max_col = sheet.ncols
        if sheet_borders[0] and sheet_borders[0] > 2:
            min_row = sheet_borders[0]
        if sheet_borders[1]:
            max_row = sheet_borders[1]
        if sheet_borders[2]:
            min_col = sheet_borders[2]
        if sheet_borders[3]:
            max_col = sheet_borders[3]

        for col_index in range(min_col, max_col):
            # Получаем значение ячейки во второй строке (индекс 1)
            cell_value = sheet.cell_value(1, col_index)
            if 'всего' in str(cell_value):
                break
            if str(cell_value) == '1':
                avoid_indexes.append(col_index)
                continue
            if cell_value is None or cell_value == '':
                avoid_indexes.append(col_index)
                continue
            if sheet.cell_type(1, col_index) == xlrd.XL_CELL_NUMBER:
                avoid_indexes.append(col_index)
                continue
            headers.append(cell_value)

        for row_index in range(min_row, max_row):
            row_values = []
            skip_row = False
            for col_index in range(len(headers)+len(avoid_indexes)):
                if col_index in avoid_indexes:
                    continue
                cell_value = sheet.cell_value(row_index, col_index)
                if 'Итого' in str(cell_value):
                    skip_row = True
                    break
                row_values.append(cell_value)
            if skip_row:
                break
            data.append(row_values)

        return headers, data
